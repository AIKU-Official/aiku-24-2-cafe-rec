from selenium.webdriver.common.by import By
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter
from openpyxl import Workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
import time
import datetime
import requests
from selenium.webdriver.chrome.service import Service
import os

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

def switch_left():
    ############## iframe으로 왼쪽 포커스 맞추기 ##############
    driver.switch_to.parent_frame()
    iframe = driver.find_element(By.XPATH,'//*[@id="searchIframe"]')
    driver.switch_to.frame(iframe)

# url
url = 'https://map.naver.com/p/search/광진구카페?c=13.00,0,0,0,dh'


# 구 별 폴더
dir = 'gwangjin'
if not os.path.exists(dir):
    os.makedirs(dir)
    
# Webdriver headless mode setting
options = webdriver.ChromeOptions()
# options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")

# BS4 setting for secondary access
session = requests.Session()
headers = {
    "User-Agent": "user value"
    }

retries = Retry(total=5,
                backoff_factor=0.1,
                status_forcelist=[500, 502, 503, 504])

session.mount('http://', HTTPAdapter(max_retries=retries))
# Chrome 드라이버 실행
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# 지정된 URL로 접속
driver.get(url)

# 페이지 로드 대기
time.sleep(5) 

store_id_list = []

# Start crawling/scraping!
# 네이버 맵에서 id 크롤링
try:
    # 엑셀 파일 생성
    store_id_xlsx = Workbook()
    list_sheet = store_id_xlsx.create_sheet('output')
    
    page_counter = 0 
    while (True): # 페이지는 5번까지만 진행
         # 페이지가 완전히 로드되도록 5초 대기
        switch_left()
        
        # 페이지 숫자를 초기에 체크 [ True / False ]
        # 이건 페이지 넘어갈때마다 계속 확인해줘야 함 (페이지 새로 로드 될때마다 버튼 상태 값이 바뀜)
        next_page = driver.find_element(By.XPATH,'//*[@id="app-root"]/div/div[2]/div[2]/a[7]').get_attribute('aria-disabled')
        if(next_page == 'true'):
            break
        
        page_no = driver.find_element(By.XPATH,'//a[contains(@class, "mBN2s qxokY")]').text
        if (int(page_no) > 5) :
            break
        print(f'현재 페이지: {page_no}')
        
        ############## 맨 밑까지 스크롤 ##############
        scrollable_element = driver.find_element(By.CLASS_NAME, "Ryr1F")
    
        last_height = driver.execute_script("return arguments[0].scrollHeight", scrollable_element)
    
        while True:
            # 요소 내에서 아래로 600px 스크롤
            driver.execute_script("arguments[0].scrollTop += 600;", scrollable_element)
    
            # 페이지 로드를 기다림
            time.sleep(5)  # 동적 콘텐츠 로드 시간에 따라 조절
    
            # 새 높이 계산
            new_height = driver.execute_script("return arguments[0].scrollHeight", scrollable_element)
    
            # 스크롤이 더 이상 늘어나지 않으면 루프 종료
            if new_height == last_height:
                break
    
            last_height = new_height
        
        ############## 현재 page number 가져오기 - 1 페이지 ##############
        # page_no = driver.find_element(By.XPATH,'//a[contains(@class, "mBN2s qxokY")]').text
        
        # <li> 요소들 찾기 - 예시로 클래스명이 'UEzoS rTjJo'인 <li>들 선택
        li_elements = driver.find_elements(By.CSS_SELECTOR, 'li.UEzoS.rTjJo')
        print(f"찾은 요소 개수: {len(li_elements)}")
        
        print('현재 ' + '\033[95m' + str(page_no) + '\033[0m' + ' 페이지 / '+ '총 ' + '\033[95m' + str(len(li_elements)) + '\033[0m' + '개의 가게를 찾았습니다.\n')

        # <li> 요소들을 하나씩 클릭
        for li in li_elements:
            try:
                # 광고 제거
                ad_element = li.find_elements(By.CSS_SELECTOR, 'span.place_blind')
                if ad_element and '광고' in ad_element[0].text:
                    print("광고 건너뛰는 중 ~.~")
                    continue 
                
                li.click()  # 요소 클릭
                time.sleep(3)  # 각 클릭 후 잠시 대기 (페이지가 로드될 시간을 고려)
                
                # 현재 URL 출력
                # ex) https://map.naver.com/p/search/%EC%84%B1%EB%B6%81%EA%B5%AC%20%EC%9D%8C%EC%8B%9D%EC%A0%90/place/1831632226?c=12.00,0,0,0,dh&placePath=%3Fentry%253Dbmp
                current_url = driver.current_url
                store_id = current_url.split('place/')[-1].split('?')[0]
                
                print(f"현재 URL: {current_url}")
                print(f"가게 ID: {store_id}")
                store_id_list.append(store_id)
                list_sheet.append([store_id])
                
            except Exception as e:
                print(f"클릭할 수 없습니다: {e}")
        
        switch_left()
        
        # 페이지 넘기고 카운터 1 증가
        driver.find_element(By.XPATH, '//*[@id="app-root"]/div/div[2]/div[2]/a[7]').click()
        page_counter += 1 # 페이지 하나 넘길 때마다 카운터 += 1
    
        # # 페이지 다음 버튼이 활성화 상태일 경우 계속 진행
        # if(next_page == 'false'):
        #     driver.find_element(By.XPATH,'//*[@id="app-root"]/div/div[2]/div[2]/a[7]').click()
        # # 아닐 경우 루프 정지
        # else:
        #     break
        # break
        
        # 일정 시간 동안 브라우저가 유지되도록 대기
        time.sleep(10)  # 10초간 창을 유지
        
    ############## 가게 ID 리스트 엑셀에 저장 ##############  
    store_id_list_path = os.path.join(dir, f'{dir}_store_id_list.xlsx')
    store_id_xlsx.save(store_id_list_path)
    
except Exception as e:
    print(f"오류 발생: {e}")

# 네이버 플레이스에서 리뷰 크롤링
try:
    now = datetime.datetime.now()
    xlsx = Workbook()
    list_sheet = xlsx.create_sheet('output')
    list_sheet.append(['store_id', 'store_name', 'address', 'menu', 'nickname', 'content', 'date', 'revisit'])

    # 각 store_id에 대해 리뷰 크롤링
    for store_id in store_id_list:
        
        # (1) home - 가게 이름, 주소
        home_url = f'https://m.place.naver.com/restaurant/{store_id}/home?entry=ple&reviewSort=recent'
        driver.get(home_url)
        driver.implicitly_wait(10)

        # 가게 이름 크롤링
        store_name = driver.find_element(By.CSS_SELECTOR, 'span.GHAhO').text

        # 가게 주소 크롤링
        address = driver.find_element(By.CSS_SELECTOR, 'span.LDgIH').text

        # (2) menu - 메뉴
        # 메뉴 페이지에서 메뉴 정보 크롤링 (가격 제외)
        menu_url = f'https://m.place.naver.com/restaurant/{store_id}/menu/list?entry=ple&reviewSort=recent'
        driver.get(menu_url)
        try:
            menu_elements = driver.find_elements(By.CSS_SELECTOR, 'span.lPzHi')
            menu = [m.text for m in menu_elements]
            menu = ', '.join(menu) if menu else 'N/A'
        except:
            menu = 'N/A' # 메뉴 등록이 안 되어 있는 경우

        print(f"가게 이름: {store_name}")
        print(f"주소: {address}")
        print(f"메뉴: {menu}")

        # (3) review - 닉네임, 리뷰 내용, 작성일, 재방문 횟수
        review_url = f'https://m.place.naver.com/restaurant/{store_id}/review/visitor?entry=ple&reviewSort=recent'
        driver.get(review_url)
        driver.implicitly_wait(10)

        # Pagedown
        driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_DOWN)

        try:
            for i in range(30):  # 최대 30번 반복
                driver.find_element(By.XPATH, '//*[@id="app-root"]/div/div/div/div[6]/div[2]/div[3]/div[2]/div/a').click()
                time.sleep(0.4)
        except Exception as e:
            print('finish')
        else :
            print('30times_finish')

        time.sleep(25)
        
        # 리뷰 크롤링
        html = driver.page_source
        bs = BeautifulSoup(html, 'lxml')
        reviews = bs.select('li.pui__X35jYm.EjjAW')

        review_count = 0
        for r in reviews:
            if review_count >= 15:  # 각 카페당 리뷰 15개까지만 크롤링
                break

            # nickname
            nickname = r.select_one('div.pui__JiVbY3 > span.pui__uslU0d')
            nickname = nickname.text if nickname else ''

            # content ('이런 점이 좋았어요' 제외)
            content = r.select_one('div.pui__vn15t2 > a.pui__xtsQN-')
            if 'pui__HLNvmI' in r.get('class', []):
                continue
            content = content.text if content else ''

            # date
            date_elements = r.select('div.pui__QKE5Pr > span.pui__gfuUIT > time')
            date = date_elements[0].text if date_elements else 'N/A'

            # revisit
            revisit_span = r.select('div.pui__QKE5Pr > span.pui__gfuUIT')
            revisit = revisit_span[1].text if len(revisit_span) > 1 else 'N/A'

            # 리스트에 추가
            list_sheet.append([store_id, store_name, address, menu, nickname, content, date, revisit])

            print(f"{nickname} / {content} / {date} / {revisit}")
            review_count += 1
            time.sleep(0.1)  # 크롤링 사이 시간 대기

    # 엑셀 파일 저장
    review_file_name = 'naver_review_' + now.strftime('%Y-%m-%d_%H-%M-%S') + '.xlsx'
    review_path = os.path.join(dir, review_file_name)
    xlsx.save(review_path)

except Exception as e:
    print(e)

finally:
    driver.quit()